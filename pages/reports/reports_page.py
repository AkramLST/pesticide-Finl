import os
from datetime import datetime, date, timedelta

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QFrame, QScrollArea, QSizePolicy, QMessageBox,
    QDateEdit, QComboBox, QTableWidget, QTableWidgetItem,
    QHeaderView, QAbstractItemView, QTabWidget, QFormLayout
)
from PySide6.QtCore import Qt, QDate
from PySide6.QtGui import QColor

from database.connection import get_connection
from utils.helpers import format_currency, format_datetime, format_date
from utils.config import EXPORTS_DIR, PAYMENT_METHODS

_GREEN = ("QPushButton{background:#2e7d32;color:white;border-radius:8px;"
          "padding:9px 20px;font-size:13px;font-weight:700;border:none;}"
          "QPushButton:hover{background:#1b5e20;}")
_AMBER = ("QPushButton{background:#fffbeb;color:#92400e;border:1px solid #fde68a;"
          "border-radius:8px;font-size:13px;font-weight:600;padding:9px 18px;}"
          "QPushButton:hover{background:#fef3c7;}")
_GRAY  = ("QPushButton{background:#f1f5f9;color:#475569;border-radius:8px;"
          "padding:9px 18px;font-size:13px;border:1.5px solid #e2e8f0;}"
          "QPushButton:hover{background:#e2e8f0;}")
_TS = """
    QTableWidget{border:none;font-size:13px;background:white;
                 alternate-background-color:#f8fafc;}
    QHeaderView::section{background:#f1f5f9;color:#475569;font-weight:700;
        font-size:12px;padding:8px;border:none;border-right:1px solid #e2e8f0;}
    QTableWidget::item{padding:8px;color:#1e293b;border-bottom:1px solid #f1f5f9;}
    QTableWidget::item:selected{background:#dbeafe;color:#1e40af;}
"""
_DE_STYLE = ("QDateEdit{border:1.5px solid #e2e8f0;border-radius:7px;"
             "padding:7px 10px;font-size:13px;background:white;}"
             "QDateEdit:focus{border-color:#2e7d32;}")


def _card(title: str) -> tuple:
    """Return (card QFrame, inner_layout QVBoxLayout)."""
    card = QFrame()
    card.setAttribute(Qt.WA_StyledBackground, True)
    card.setStyleSheet(
        "QFrame{background:white;border-radius:12px;border:1px solid #e2e8f0;}")
    card.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Minimum)
    card.setContentsMargins(0, 0, 0, 0)
    layout = QVBoxLayout(card)
    layout.setContentsMargins(20, 16, 20, 16)
    layout.setSpacing(12)
    hdr = QLabel(title)
    hdr.setStyleSheet(
        "font-size:15px;font-weight:700;color:#0f172a;"
        "padding-bottom:4px;border-bottom:2px solid #2e7d32;")
    layout.addWidget(hdr)
    return card, layout


def _table(cols: list) -> QTableWidget:
    t = QTableWidget()
    t.setColumnCount(len(cols))
    t.setHorizontalHeaderLabels(cols)
    t.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
    t.verticalHeader().setVisible(False)
    t.setEditTriggers(QAbstractItemView.NoEditTriggers)
    t.setSelectionBehavior(QAbstractItemView.SelectRows)
    t.setAlternatingRowColors(True)
    t.setShowGrid(False)
    t.setStyleSheet(_TS)
    return t


def _fill(table: QTableWidget, rows: list, color_col: int = -1,
          color_fn=None):
    table.setSortingEnabled(False)
    table.setRowCount(len(rows))
    for r, row in enumerate(rows):
        for c, val in enumerate(row):
            item = QTableWidgetItem(str(val))
            item.setTextAlignment(Qt.AlignVCenter | Qt.AlignLeft)
            if color_col == c and color_fn:
                item.setForeground(QColor(color_fn(row)))
            table.setItem(r, c, item)
        table.setRowHeight(r, 40)
    table.setSortingEnabled(True)


# ─────────────────────────────────────────────────────────────────────────────
class ReportsPage(QWidget):

    def __init__(self):
        super().__init__()
        self.setAttribute(Qt.WA_StyledBackground, True)
        self.setStyleSheet("background:#f1f5f9;")
        self._build_ui()

    def _build_ui(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        # Page header
        header = QFrame()
        header.setFixedHeight(64)
        header.setAttribute(Qt.WA_StyledBackground, True)
        header.setStyleSheet(
            "QFrame{background:white;border-bottom:1.5px solid #e2e8f0;}")
        hl = QHBoxLayout(header)
        hl.setContentsMargins(24, 0, 24, 0)
        hl.addWidget(QLabel("📊  Reports",
            styleSheet="font-size:20px;font-weight:700;color:#0f172a;"))
        hl.addStretch()
        outer.addWidget(header)

        # Tabs
        tabs = QTabWidget()
        tabs.setStyleSheet("""
            QTabWidget::pane{border:none;background:#f1f5f9;}
            QTabBar::tab{background:#f1f5f9;color:#475569;padding:10px 20px;
                font-size:13px;font-weight:600;border:none;
                border-bottom:3px solid transparent;}
            QTabBar::tab:selected{color:#2e7d32;
                border-bottom:3px solid #2e7d32;background:white;}
            QTabBar::tab:hover{background:#e2e8f0;}
        """)

        tabs.addTab(self._build_sales_tab(),    "📈  Sales")
        tabs.addTab(self._build_profit_tab(),   "💰  Profit / Loss")
        tabs.addTab(self._build_dues_tab(),     "⏳  Customer Dues")
        tabs.addTab(self._build_inventory_tab(),"📦  Inventory")
        outer.addWidget(tabs)

    # ── SALES REPORT ─────────────────────────────────────────────────────────
    def _build_sales_tab(self):
        w = QScrollArea()
        w.setWidgetResizable(True)
        w.setFrameShape(QFrame.NoFrame)
        w.setStyleSheet("background:#f1f5f9;")
        body = QWidget(); body.setStyleSheet("background:#f1f5f9;")
        bl = QVBoxLayout(body)
        bl.setContentsMargins(24, 20, 24, 24)
        bl.setSpacing(18)

        # Filter bar
        fcard, fl = _card("🔎  Filters")
        frow = QHBoxLayout()
        self.s_from = QDateEdit(QDate.currentDate().addDays(-30))
        self.s_to   = QDateEdit(QDate.currentDate())
        for de in (self.s_from, self.s_to):
            de.setCalendarPopup(True)
            de.setStyleSheet(_DE_STYLE)
            de.setFixedWidth(140)
        self.s_method = QComboBox()
        self.s_method.addItem("All Methods")
        self.s_method.addItems(PAYMENT_METHODS)
        self.s_method.setStyleSheet(
            "QComboBox{border:1.5px solid #e2e8f0;border-radius:7px;"
            "padding:7px 10px;font-size:13px;background:white;}")
        self.s_method.setFixedWidth(150)
        gen_btn = QPushButton("▶  Generate")
        gen_btn.setStyleSheet(_GREEN)
        gen_btn.setFixedHeight(38)
        gen_btn.clicked.connect(self._gen_sales)
        pdf_btn = QPushButton("📄  PDF")
        pdf_btn.setStyleSheet(_AMBER)
        pdf_btn.setFixedHeight(38)
        pdf_btn.clicked.connect(self._export_sales_pdf)
        xl_btn = QPushButton("📊  Excel")
        xl_btn.setStyleSheet(_GRAY)
        xl_btn.setFixedHeight(38)
        xl_btn.clicked.connect(self._export_sales_excel)

        frow.addWidget(QLabel("From:", styleSheet="font-size:13px;"))
        frow.addWidget(self.s_from)
        frow.addWidget(QLabel("To:", styleSheet="font-size:13px;"))
        frow.addWidget(self.s_to)
        frow.addWidget(QLabel("Method:", styleSheet="font-size:13px;"))
        frow.addWidget(self.s_method)
        frow.addStretch()
        frow.addWidget(gen_btn)
        frow.addWidget(xl_btn)
        frow.addWidget(pdf_btn)
        fl.addLayout(frow)
        bl.addWidget(fcard)

        # Summary strip
        self._sales_summary = QLabel("—")
        self._sales_summary.setStyleSheet(
            "font-size:13px;color:#475569;padding:8px 0;")
        bl.addWidget(self._sales_summary)

        # Table
        tcard, tl = _card("📋  Sales Records")
        self._sales_table = _table([
            "Invoice #", "Customer", "Product", "Qty",
            "Total", "Paid", "Remaining", "Method", "Date", "Sold By"
        ])
        self._sales_table.setMinimumHeight(350)
        tl.addWidget(self._sales_table)
        bl.addWidget(tcard)
        bl.addStretch()
        w.setWidget(body)
        return w

    def _sales_query(self) -> list:
        d_from  = self.s_from.date().toString("yyyy-MM-dd")
        d_to    = self.s_to.date().toString("yyyy-MM-dd") + " 23:59:59"
        method  = self.s_method.currentText()
        conn = get_connection()
        q = """
            SELECT s.id, s.invoice_number,
                   COALESCE(c.name,'Walk-in') AS customer_name,
                   s.total_amount, s.paid_amount, s.remaining_amount,
                   s.payment_method, s.sale_date,
                   COALESCE(u.name, u.username, 'Unknown') AS seller_name,
                   s.discount
            FROM sales s
            LEFT JOIN customers c ON s.customer_id = c.id
            LEFT JOIN users u     ON s.sold_by = u.id
            WHERE s.is_deleted=0
              AND s.sale_date BETWEEN ? AND ?
        """
        params = [d_from, d_to]
        if method != "All Methods":
            q += " AND s.payment_method=?"
            params.append(method)
        q += " ORDER BY s.sale_date DESC"
        rows = conn.execute(q, params).fetchall()
        # attach first product name
        result = []
        for r in rows:
            items = conn.execute(
                "SELECT p.name, si.quantity FROM sale_items si "
                "LEFT JOIN products p ON si.product_id=p.id "
                "WHERE si.sale_id=?", (r["id"],)
            ).fetchall()
            prod = items[0]["name"] if items else "—"
            qty  = sum(i["quantity"] for i in items)
            result.append({**dict(r), "_product": prod, "_qty": qty})
        conn.close()
        return result

    def _gen_sales(self):
        self._sales_data = self._sales_query()
        rows = [
            [s["invoice_number"], s["customer_name"], s["_product"],
             s["_qty"], format_currency(s["total_amount"]),
             format_currency(s["paid_amount"]),
             format_currency(s["remaining_amount"]),
             s["payment_method"] or "—",
             format_datetime(s["sale_date"]),
             s["seller_name"]]
            for s in self._sales_data
        ]
        _fill(self._sales_table, rows, color_col=6,
              color_fn=lambda r: "#dc2626" if "." in r[6] and
              float(r[6].replace("Rs","").replace(",","").strip()) > 0
              else "#16a34a")
        total   = sum(s["total_amount"] for s in self._sales_data)
        paid    = sum(s["paid_amount"]  for s in self._sales_data)
        pending = sum(s["remaining_amount"] for s in self._sales_data)
        self._sales_summary.setText(
            f"<b>Records:</b> {len(self._sales_data)}  │  "
            f"<b>Revenue:</b> {format_currency(total)}  │  "
            f"<b>Collected:</b> {format_currency(paid)}  │  "
            f"<b>Pending:</b> <span style='color:#dc2626'>"
            f"{format_currency(pending)}</span>")

    def _export_sales_pdf(self):
        if not getattr(self, "_sales_data", None):
            self._gen_sales()
        if not self._sales_data:
            return QMessageBox.information(self, "Empty", "No data.")
        try:
            from services.export_service import export_sales_pdf
            path = export_sales_pdf(self._sales_data)
            QMessageBox.information(self, "Exported", f"PDF:\n{path}")
            os.startfile(path)
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    def _export_sales_excel(self):
        if not getattr(self, "_sales_data", None):
            self._gen_sales()
        if not self._sales_data:
            return QMessageBox.information(self, "Empty", "No data.")
        try:
            from services.export_service import export_sales_excel
            path = export_sales_excel(self._sales_data)
            QMessageBox.information(self, "Exported", f"Excel:\n{path}")
            os.startfile(path)
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    # ── PROFIT / LOSS REPORT ─────────────────────────────────────────────────
    def _build_profit_tab(self):
        w = QScrollArea(); w.setWidgetResizable(True)
        w.setFrameShape(QFrame.NoFrame); w.setStyleSheet("background:#f1f5f9;")
        body = QWidget(); body.setStyleSheet("background:#f1f5f9;")
        bl = QVBoxLayout(body)
        bl.setContentsMargins(24, 20, 24, 24); bl.setSpacing(18)

        fcard, fl = _card("🔎  Filters")
        frow = QHBoxLayout()
        self.p_from = QDateEdit(QDate.currentDate().addDays(-30))
        self.p_to   = QDateEdit(QDate.currentDate())
        for de in (self.p_from, self.p_to):
            de.setCalendarPopup(True); de.setStyleSheet(_DE_STYLE)
            de.setFixedWidth(140)
        gen_btn = QPushButton("▶  Generate"); gen_btn.setStyleSheet(_GREEN)
        gen_btn.setFixedHeight(38); gen_btn.clicked.connect(self._gen_profit)
        pdf_btn = QPushButton("📄  PDF"); pdf_btn.setStyleSheet(_AMBER)
        pdf_btn.setFixedHeight(38); pdf_btn.clicked.connect(self._export_profit_pdf)
        frow.addWidget(QLabel("From:", styleSheet="font-size:13px;"))
        frow.addWidget(self.p_from)
        frow.addWidget(QLabel("To:", styleSheet="font-size:13px;"))
        frow.addWidget(self.p_to)
        frow.addStretch(); frow.addWidget(gen_btn); frow.addWidget(pdf_btn)
        fl.addLayout(frow); bl.addWidget(fcard)

        self._profit_summary = QLabel("—")
        self._profit_summary.setStyleSheet("font-size:13px;color:#475569;padding:8px 0;")
        bl.addWidget(self._profit_summary)

        tcard, tl = _card("💰  Profit by Product")
        self._profit_table = _table([
            "Product", "Category", "Units Sold",
            "Revenue", "Cost", "Gross Profit", "Margin %"
        ])
        self._profit_table.setMinimumHeight(350)
        tl.addWidget(self._profit_table)
        bl.addWidget(tcard); bl.addStretch()
        w.setWidget(body); return w

    def _gen_profit(self):
        d_from = self.p_from.date().toString("yyyy-MM-dd")
        d_to   = self.p_to.date().toString("yyyy-MM-dd") + " 23:59:59"
        conn = get_connection()
        rows = conn.execute("""
            SELECT p.name, p.category,
                   SUM(si.quantity) AS units_sold,
                   SUM(si.subtotal) AS revenue,
                   SUM(si.quantity * p.purchase_price) AS cost
            FROM sale_items si
            JOIN products p ON si.product_id = p.id
            JOIN sales s    ON si.sale_id = s.id
            WHERE s.is_deleted=0 AND s.sale_date BETWEEN ? AND ?
            GROUP BY p.id ORDER BY revenue DESC
        """, (d_from, d_to)).fetchall()
        conn.close()
        self._profit_data = [dict(r) for r in rows]

        table_rows = []
        total_rev = total_cost = 0.0
        for r in self._profit_data:
            rev  = r["revenue"] or 0
            cost = r["cost"] or 0
            gp   = rev - cost
            margin = f"{(gp/rev*100):.1f}%" if rev else "—"
            total_rev  += rev
            total_cost += cost
            table_rows.append([
                r["name"], r["category"] or "—", str(r["units_sold"]),
                format_currency(rev), format_currency(cost),
                format_currency(gp), margin
            ])

        _fill(self._profit_table, table_rows)
        gross = total_rev - total_cost
        color = "#16a34a" if gross >= 0 else "#dc2626"
        self._profit_summary.setText(
            f"<b>Revenue:</b> {format_currency(total_rev)}  │  "
            f"<b>Cost:</b> {format_currency(total_cost)}  │  "
            f"<b>Gross Profit:</b> <span style='color:{color}'>"
            f"{format_currency(gross)}</span>")

    def _export_profit_pdf(self):
        if not getattr(self, "_profit_data", None):
            self._gen_profit()
        if not self._profit_data:
            return QMessageBox.information(self, "Empty", "No data.")
        try:
            path = self._write_profit_pdf()
            QMessageBox.information(self, "Exported", f"PDF:\n{path}")
            os.startfile(path)
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    def _write_profit_pdf(self) -> str:
        import os
        from datetime import datetime
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                        Paragraph, Spacer, HRFlowable)
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        os.makedirs(EXPORTS_DIR, exist_ok=True)
        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = os.path.join(EXPORTS_DIR, f"profit_{ts}.pdf")
        doc  = SimpleDocTemplate(path, pagesize=landscape(A4),
                                 rightMargin=1.2*cm, leftMargin=1.2*cm,
                                 topMargin=1.5*cm, bottomMargin=1.2*cm)
        styles = getSampleStyleSheet()
        title_s = ParagraphStyle("t", parent=styles["Normal"], fontSize=14,
                                 fontName="Helvetica-Bold",
                                 textColor=colors.HexColor("#0f172a"),
                                 spaceAfter=4)
        sub_s   = ParagraphStyle("s", parent=styles["Normal"], fontSize=9,
                                 textColor=colors.HexColor("#64748b"),
                                 spaceAfter=8)
        story = [
            Paragraph("Profit / Loss Report", title_s),
            Paragraph(f"Period: {self.p_from.date().toString('dd-MM-yyyy')} "
                      f"to {self.p_to.date().toString('dd-MM-yyyy')}  |  "
                      f"Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}",
                      sub_s),
            HRFlowable(width="100%", thickness=1,
                       color=colors.HexColor("#2e7d32"), spaceAfter=8),
        ]
        hdrs = ["Product","Category","Units Sold",
                "Revenue","Cost","Gross Profit","Margin %"]
        data = [hdrs]
        for r in self._profit_data:
            rev  = r["revenue"] or 0
            cost = r["cost"] or 0
            gp   = rev - cost
            margin = f"{(gp/rev*100):.1f}%" if rev else "—"
            data.append([r["name"], r["category"] or "—",
                         str(r["units_sold"]),
                         format_currency(rev), format_currency(cost),
                         format_currency(gp), margin])
        col_w = [6*cm,3*cm,2.5*cm,4*cm,4*cm,4*cm,3*cm]
        tbl = Table(data, colWidths=col_w, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#2e7d32")),
            ("TEXTCOLOR", (0,0),(-1,0),colors.white),
            ("FONTNAME",  (0,0),(-1,0),"Helvetica-Bold"),
            ("FONTSIZE",  (0,0),(-1,-1),9),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),
             [colors.white, colors.HexColor("#f8fafc")]),
            ("GRID",      (0,0),(-1,-1),0.4,colors.HexColor("#e2e8f0")),
            ("TOPPADDING",(0,0),(-1,-1),5),
            ("BOTTOMPADDING",(0,0),(-1,-1),5),
        ]))
        story.append(tbl)
        doc.build(story)
        return path

    # ── CUSTOMER DUES REPORT ─────────────────────────────────────────────────
    def _build_dues_tab(self):
        w = QScrollArea(); w.setWidgetResizable(True)
        w.setFrameShape(QFrame.NoFrame); w.setStyleSheet("background:#f1f5f9;")
        body = QWidget(); body.setStyleSheet("background:#f1f5f9;")
        bl = QVBoxLayout(body)
        bl.setContentsMargins(24, 20, 24, 24); bl.setSpacing(18)

        fcard, fl = _card("🔎  Options")
        frow = QHBoxLayout()
        gen_btn = QPushButton("▶  Load Dues"); gen_btn.setStyleSheet(_GREEN)
        gen_btn.setFixedHeight(38); gen_btn.clicked.connect(self._gen_dues)
        pdf_btn = QPushButton("📄  Export PDF"); pdf_btn.setStyleSheet(_AMBER)
        pdf_btn.setFixedHeight(38); pdf_btn.clicked.connect(self._export_dues_pdf)
        xl_btn  = QPushButton("📊  Export Excel"); xl_btn.setStyleSheet(_GRAY)
        xl_btn.setFixedHeight(38);  xl_btn.clicked.connect(self._export_dues_excel)
        frow.addWidget(gen_btn); frow.addWidget(xl_btn); frow.addWidget(pdf_btn)
        frow.addStretch(); fl.addLayout(frow); bl.addWidget(fcard)

        self._dues_summary = QLabel("—")
        self._dues_summary.setStyleSheet(
            "font-size:13px;color:#475569;padding:8px 0;")
        bl.addWidget(self._dues_summary)

        tcard, tl = _card("⏳  Customers with Pending Balance")
        self._dues_table = _table([
            "ID", "Name", "Phone", "Total Paid",
            "Total Pending", "Last Purchase"
        ])
        self._dues_table.setMinimumHeight(350)
        tl.addWidget(self._dues_table)
        bl.addWidget(tcard); bl.addStretch()
        w.setWidget(body); return w

    def _gen_dues(self):
        conn = get_connection()
        rows = conn.execute("""
            SELECT id, name, phone, total_paid, total_pending, last_purchase_date
            FROM customers WHERE is_active=1 AND total_pending > 0
            ORDER BY total_pending DESC
        """).fetchall()
        conn.close()
        self._dues_data = [dict(r) for r in rows]
        table_rows = [
            [str(r["id"]), r["name"], r["phone"] or "—",
             format_currency(r["total_paid"]),
             format_currency(r["total_pending"]),
             format_datetime(r["last_purchase_date"] or "")]
            for r in self._dues_data
        ]
        _fill(self._dues_table, table_rows, color_col=4,
              color_fn=lambda _: "#dc2626")
        total_due = sum(r["total_pending"] for r in self._dues_data)
        self._dues_summary.setText(
            f"<b>Customers with dues:</b> {len(self._dues_data)}  │  "
            f"<b>Total Outstanding:</b> "
            f"<span style='color:#dc2626'>{format_currency(total_due)}</span>")

    def _export_dues_pdf(self):
        if not getattr(self, "_dues_data", None):
            self._gen_dues()
        if not self._dues_data:
            return QMessageBox.information(self, "Empty", "No dues.")
        try:
            import os, openpyxl
            from datetime import datetime
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.units import cm
            from reportlab.platypus import (SimpleDocTemplate, Table,
                                            TableStyle, Paragraph,
                                            Spacer, HRFlowable)
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            os.makedirs(EXPORTS_DIR, exist_ok=True)
            ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
            path = os.path.join(EXPORTS_DIR, f"customer_dues_{ts}.pdf")
            doc  = SimpleDocTemplate(path, pagesize=A4,
                                     rightMargin=1.5*cm, leftMargin=1.5*cm,
                                     topMargin=1.5*cm, bottomMargin=1.5*cm)
            styles = getSampleStyleSheet()
            title_s = ParagraphStyle("t", parent=styles["Normal"], fontSize=14,
                                     fontName="Helvetica-Bold", spaceAfter=4)
            sub_s   = ParagraphStyle("s", parent=styles["Normal"], fontSize=9,
                                     textColor=colors.HexColor("#64748b"),
                                     spaceAfter=8)
            total_due = sum(r["total_pending"] for r in self._dues_data)
            story = [
                Paragraph("Customer Dues Report", title_s),
                Paragraph(f"Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}  |  "
                          f"Total outstanding: {format_currency(total_due)}", sub_s),
                HRFlowable(width="100%", thickness=1,
                           color=colors.HexColor("#2e7d32"), spaceAfter=8),
            ]
            hdrs = ["ID","Name","Phone","Paid","Pending","Last Purchase"]
            data = [hdrs] + [
                [str(r["id"]), r["name"], r["phone"] or "—",
                 format_currency(r["total_paid"]),
                 format_currency(r["total_pending"]),
                 format_datetime(r["last_purchase_date"] or "")]
                for r in self._dues_data
            ]
            col_w = [1.5*cm,5*cm,3.5*cm,4*cm,4*cm,4*cm]
            tbl = Table(data, colWidths=col_w, repeatRows=1)
            tbl.setStyle(TableStyle([
                ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#2e7d32")),
                ("TEXTCOLOR", (0,0),(-1,0),colors.white),
                ("FONTNAME",  (0,0),(-1,0),"Helvetica-Bold"),
                ("FONTSIZE",  (0,0),(-1,-1),9),
                ("ROWBACKGROUNDS",(0,1),(-1,-1),
                 [colors.white,colors.HexColor("#f8fafc")]),
                ("GRID",(0,0),(-1,-1),0.4,colors.HexColor("#e2e8f0")),
                ("TOPPADDING",(0,0),(-1,-1),5),
                ("BOTTOMPADDING",(0,0),(-1,-1),5),
                ("TEXTCOLOR",(4,1),(-1,-1),colors.HexColor("#dc2626")),
            ]))
            story.append(tbl)
            doc.build(story)
            QMessageBox.information(self, "Exported", f"PDF:\n{path}")
            os.startfile(path)
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    def _export_dues_excel(self):
        if not getattr(self, "_dues_data", None):
            self._gen_dues()
        if not self._dues_data:
            return QMessageBox.information(self, "Empty", "No dues.")
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            os.makedirs(EXPORTS_DIR, exist_ok=True)
            ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
            path = os.path.join(EXPORTS_DIR, f"customer_dues_{ts}.xlsx")
            wb = openpyxl.Workbook(); ws = wb.active
            ws.title = "Customer Dues"
            hdrs = ["ID","Name","Phone","Total Paid","Total Pending","Last Purchase"]
            for c, h in enumerate(hdrs, 1):
                cell = ws.cell(1, c, h)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="2e7d32")
                cell.alignment = Alignment(horizontal="center")
            for i, r in enumerate(self._dues_data, 2):
                ws.append([
                    r["id"], r["name"], r["phone"] or "—",
                    r["total_paid"], r["total_pending"],
                    format_datetime(r["last_purchase_date"] or "")
                ])
            for col in ws.columns:
                from openpyxl.utils import get_column_letter
                ws.column_dimensions[get_column_letter(
                    col[0].column)].width = 20
            wb.save(path)
            QMessageBox.information(self, "Exported", f"Excel:\n{path}")
            os.startfile(path)
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    # ── INVENTORY STATUS REPORT ───────────────────────────────────────────────
    def _build_inventory_tab(self):
        w = QScrollArea(); w.setWidgetResizable(True)
        w.setFrameShape(QFrame.NoFrame); w.setStyleSheet("background:#f1f5f9;")
        body = QWidget(); body.setStyleSheet("background:#f1f5f9;")
        bl = QVBoxLayout(body)
        bl.setContentsMargins(24, 20, 24, 24); bl.setSpacing(18)

        fcard, fl = _card("🔎  Options")
        frow = QHBoxLayout()
        self.inv_filter = QComboBox()
        self.inv_filter.addItems(["All Products", "Low Stock Only",
                                  "Out of Stock Only", "Expired Only"])
        self.inv_filter.setStyleSheet(
            "QComboBox{border:1.5px solid #e2e8f0;border-radius:7px;"
            "padding:7px 10px;font-size:13px;background:white;}")
        self.inv_filter.setFixedWidth(200)
        gen_btn = QPushButton("▶  Load"); gen_btn.setStyleSheet(_GREEN)
        gen_btn.setFixedHeight(38); gen_btn.clicked.connect(self._gen_inventory)
        pdf_btn = QPushButton("📄  PDF"); pdf_btn.setStyleSheet(_AMBER)
        pdf_btn.setFixedHeight(38); pdf_btn.clicked.connect(self._export_inv_pdf)
        xl_btn  = QPushButton("📊  Excel"); xl_btn.setStyleSheet(_GRAY)
        xl_btn.setFixedHeight(38);  xl_btn.clicked.connect(self._export_inv_excel)
        frow.addWidget(self.inv_filter)
        frow.addStretch()
        frow.addWidget(gen_btn); frow.addWidget(xl_btn); frow.addWidget(pdf_btn)
        fl.addLayout(frow); bl.addWidget(fcard)

        self._inv_summary = QLabel("—")
        self._inv_summary.setStyleSheet(
            "font-size:13px;color:#475569;padding:8px 0;")
        bl.addWidget(self._inv_summary)

        tcard, tl = _card("📦  Inventory Snapshot")
        self._inv_table = _table([
            "ID","Name","Category","Qty","Buy Price","Sale Price","Expiry","Status"
        ])
        self._inv_table.setMinimumHeight(350)
        tl.addWidget(self._inv_table)
        bl.addWidget(tcard); bl.addStretch()
        w.setWidget(body); return w

    def _gen_inventory(self):
        from utils.helpers import is_expired, is_expiring_soon
        flt = self.inv_filter.currentText()
        conn = get_connection()
        rows = conn.execute(
            "SELECT id,name,category,quantity,low_stock_threshold,"
            "purchase_price,sale_price,expiry_date FROM products "
            "WHERE is_active=1 AND COALESCE(secret_product,0)=0"
        ).fetchall()
        conn.close()
        self._inv_data = []
        for r in rows:
            qty = r["quantity"]; thr = r["low_stock_threshold"] or 5
            exp = r["expiry_date"] or ""
            if is_expired(exp):      status = "Expired"
            elif qty == 0:           status = "Out of Stock"
            elif qty <= thr:         status = "Low Stock"
            else:                    status = "In Stock"
            if flt == "Low Stock Only"     and status != "Low Stock":    continue
            if flt == "Out of Stock Only"  and status != "Out of Stock": continue
            if flt == "Expired Only"       and status != "Expired":      continue
            self._inv_data.append({**dict(r), "_status": status})

        table_rows = [
            [str(r["id"]), r["name"], r["category"] or "—",
             str(r["quantity"]),
             format_currency(r["purchase_price"]),
             format_currency(r["sale_price"]),
             r["expiry_date"] or "—", r["_status"]]
            for r in self._inv_data
        ]
        _fill(self._inv_table, table_rows)
        self._inv_summary.setText(
            f"<b>Showing:</b> {len(self._inv_data)} products")

    def _export_inv_pdf(self):
        if not getattr(self, "_inv_data", None):
            self._gen_inventory()
        if not self._inv_data:
            return QMessageBox.information(self, "Empty", "No data.")
        try:
            from services.export_service import export_inventory_pdf
            path = export_inventory_pdf(self._inv_data)
            QMessageBox.information(self, "Exported", f"PDF:\n{path}")
            os.startfile(path)
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    def _export_inv_excel(self):
        if not getattr(self, "_inv_data", None):
            self._gen_inventory()
        if not self._inv_data:
            return QMessageBox.information(self, "Empty", "No data.")
        try:
            from services.export_service import export_inventory_excel
            path = export_inventory_excel(self._inv_data)
            QMessageBox.information(self, "Exported", f"Excel:\n{path}")
            os.startfile(path)
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))
