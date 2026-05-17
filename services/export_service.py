"""
Shared export helpers: Excel (openpyxl) and PDF (reportlab) for
inventory, sales, and customer statement reports.
"""
import os
from datetime import datetime

# ── openpyxl ─────────────────────────────────────────────────────────────────
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── reportlab ─────────────────────────────────────────────────────────────────
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, HRFlowable
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

from utils.config import EXPORTS_DIR

_H_FILL  = "2e7d32"
_H_FONT  = "FFFFFF"
_ALT     = "f8fafc"
_BORDER  = Side(style="thin", color="e2e8f0")


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL helpers
# ─────────────────────────────────────────────────────────────────────────────

def _xl_header_style(ws, row: int, cols: int):
    fill = PatternFill("solid", fgColor=_H_FILL)
    font = Font(bold=True, color=_H_FONT, size=11)
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _xl_row_style(ws, row: int, cols: int, alt: bool):
    fill = PatternFill("solid", fgColor=_ALT) if alt else None
    bd   = Border(left=_BORDER, right=_BORDER,
                  top=_BORDER, bottom=_BORDER)
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        if fill:
            cell.fill = fill
        cell.border = bd
        cell.alignment = Alignment(vertical="center")


def _xl_auto_width(ws):
    for col in ws.columns:
        max_w = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_w = max(max_w, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_w + 4, 40)


# ─────────────────────────────────────────────────────────────────────────────
# PDF helpers
# ─────────────────────────────────────────────────────────────────────────────

def _pdf_table_style(col_count: int) -> TableStyle:
    return TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0),  colors.HexColor("#2e7d32")),
        ("TEXTCOLOR",     (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0),  10),
        ("ALIGN",         (0, 0), (-1, -1), "LEFT"),
        ("FONTSIZE",      (0, 1), (-1, -1), 9),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1),
         [colors.white, colors.HexColor("#f8fafc")]),
        ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#e2e8f0")),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING",   (0, 0), (-1, -1), 6),
    ])


def _pdf_doc(filename: str, landscape_mode: bool = False):
    size = landscape(A4) if landscape_mode else A4
    return SimpleDocTemplate(
        filename, pagesize=size,
        rightMargin=1.2 * cm, leftMargin=1.2 * cm,
        topMargin=1.5 * cm, bottomMargin=1.2 * cm,
    )


def _pdf_title_para(text: str):
    styles = getSampleStyleSheet()
    s = ParagraphStyle("ttl", parent=styles["Normal"],
                       fontSize=14, fontName="Helvetica-Bold",
                       textColor=colors.HexColor("#0f172a"),
                       spaceAfter=4)
    return Paragraph(text, s)


def _pdf_sub_para(text: str):
    styles = getSampleStyleSheet()
    s = ParagraphStyle("sub", parent=styles["Normal"],
                       fontSize=9, textColor=colors.HexColor("#64748b"),
                       spaceAfter=8)
    return Paragraph(text, s)


# ─────────────────────────────────────────────────────────────────────────────
# INVENTORY EXPORT
# ─────────────────────────────────────────────────────────────────────────────

_INV_HEADERS = [
    "ID", "Product Name", "Category", "Supplier",
    "Qty", "Buy Price", "Sale Price", "Expiry Date", "Status"
]


def _inv_row(p: dict) -> list:
    from utils.helpers import is_expired, is_expiring_soon, format_currency
    qty = p.get("quantity", 0)
    thr = p.get("low_stock_threshold", 5)
    exp = p.get("expiry_date", "") or ""
    if is_expired(exp):      status = "Expired"
    elif qty == 0:           status = "Out of Stock"
    elif qty <= thr:         status = "Low Stock"
    else:                    status = "In Stock"
    return [
        p.get("id", ""),
        p.get("name", ""),
        p.get("category", ""),
        p.get("supplier_name") or "—",
        qty,
        format_currency(p.get("purchase_price", 0)),
        format_currency(p.get("sale_price", 0)),
        exp or "—",
        status,
    ]


def export_inventory_excel(products: list) -> str:
    os.makedirs(EXPORTS_DIR, exist_ok=True)
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(EXPORTS_DIR, f"inventory_{ts}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"
    ws.row_dimensions[1].height = 22

    for col, h in enumerate(_INV_HEADERS, 1):
        ws.cell(row=1, column=col, value=h)
    _xl_header_style(ws, 1, len(_INV_HEADERS))

    for i, p in enumerate(products, 2):
        for col, val in enumerate(_inv_row(p), 1):
            ws.cell(row=i, column=col, value=val)
        _xl_row_style(ws, i, len(_INV_HEADERS), i % 2 == 0)

    _xl_auto_width(ws)
    wb.save(path)
    return path


def export_inventory_pdf(products: list) -> str:
    os.makedirs(EXPORTS_DIR, exist_ok=True)
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(EXPORTS_DIR, f"inventory_{ts}.pdf")
    doc  = _pdf_doc(path, landscape_mode=True)

    story = [
        _pdf_title_para("Inventory Report"),
        _pdf_sub_para(f"Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}  |  "
                      f"Total Products: {len(products)}"),
        HRFlowable(width="100%", thickness=1,
                   color=colors.HexColor("#2e7d32"), spaceAfter=8),
    ]

    data = [_INV_HEADERS] + [_inv_row(p) for p in products]
    col_w = [1*cm, 5.5*cm, 3*cm, 3.5*cm, 1.5*cm, 3*cm, 3*cm, 3*cm, 3*cm]
    tbl  = Table(data, colWidths=col_w, repeatRows=1)
    tbl.setStyle(_pdf_table_style(len(_INV_HEADERS)))
    story.append(tbl)
    doc.build(story)
    return path


# ─────────────────────────────────────────────────────────────────────────────
# SALES EXPORT
# ─────────────────────────────────────────────────────────────────────────────

_SALE_HEADERS = [
    "Invoice #", "Customer", "Product", "Qty",
    "Discount", "Total", "Paid", "Remaining", "Method", "Date", "Sold By"
]


def _sale_row(s: dict) -> list:
    from utils.helpers import format_currency, format_datetime
    return [
        s.get("invoice_number", ""),
        s.get("customer_name") or "Walk-in",
        s.get("_product", ""),
        s.get("_qty", ""),
        format_currency(s.get("discount", 0) or 0),
        format_currency(s.get("total_amount", 0) or 0),
        format_currency(s.get("paid_amount", 0) or 0),
        format_currency(s.get("remaining_amount", 0) or 0),
        s.get("payment_method", ""),
        format_datetime(s.get("sale_date", "") or ""),
        s.get("seller_name") or "—",
    ]


def export_sales_excel(sales: list) -> str:
    os.makedirs(EXPORTS_DIR, exist_ok=True)
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(EXPORTS_DIR, f"sales_{ts}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws.row_dimensions[1].height = 22

    for col, h in enumerate(_SALE_HEADERS, 1):
        ws.cell(row=1, column=col, value=h)
    _xl_header_style(ws, 1, len(_SALE_HEADERS))

    for i, s in enumerate(sales, 2):
        for col, val in enumerate(_sale_row(s), 1):
            ws.cell(row=i, column=col, value=val)
        _xl_row_style(ws, i, len(_SALE_HEADERS), i % 2 == 0)

    _xl_auto_width(ws)
    wb.save(path)
    return path


def export_sales_pdf(sales: list) -> str:
    os.makedirs(EXPORTS_DIR, exist_ok=True)
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(EXPORTS_DIR, f"sales_{ts}.pdf")
    doc  = _pdf_doc(path, landscape_mode=True)

    from utils.helpers import format_currency
    total_rev  = sum(s.get("total_amount", 0) or 0 for s in sales)
    total_pend = sum(s.get("remaining_amount", 0) or 0 for s in sales)

    story = [
        _pdf_title_para("Sales Report"),
        _pdf_sub_para(
            f"Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}  |  "
            f"Records: {len(sales)}  |  "
            f"Total Revenue: {format_currency(total_rev)}  |  "
            f"Pending: {format_currency(total_pend)}"
        ),
        HRFlowable(width="100%", thickness=1,
                   color=colors.HexColor("#2e7d32"), spaceAfter=8),
    ]

    data = [_SALE_HEADERS] + [_sale_row(s) for s in sales]
    col_w = [2.8*cm, 3.5*cm, 4*cm, 1.2*cm,
             2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 3.5*cm, 2.5*cm]
    tbl  = Table(data, colWidths=col_w, repeatRows=1)
    tbl.setStyle(_pdf_table_style(len(_SALE_HEADERS)))
    story.append(tbl)
    doc.build(story)
    return path


# ─────────────────────────────────────────────────────────────────────────────
# CUSTOMER STATEMENT PDF
# ─────────────────────────────────────────────────────────────────────────────

def export_customer_statement(customer: dict, sales: list) -> str:
    os.makedirs(EXPORTS_DIR, exist_ok=True)
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    name = customer.get("name", "customer").replace(" ", "_")
    path = os.path.join(EXPORTS_DIR, f"statement_{name}_{ts}.pdf")
    doc  = _pdf_doc(path)

    from utils.helpers import format_currency, format_datetime
    from models.settings_model import get_setting
    shop_name = get_setting("shop_name") or "Jadeed Zarai Markaz"

    styles  = getSampleStyleSheet()
    h1_s    = ParagraphStyle("h1", parent=styles["Normal"],
                             fontSize=18, fontName="Helvetica-Bold",
                             textColor=colors.HexColor("#1b5e20"),
                             alignment=TA_CENTER, spaceAfter=2)
    sub_s   = ParagraphStyle("sub", parent=styles["Normal"],
                             fontSize=9, textColor=colors.HexColor("#64748b"),
                             alignment=TA_CENTER, spaceAfter=4)
    lbl_s   = ParagraphStyle("lbl", parent=styles["Normal"],
                             fontSize=10, fontName="Helvetica-Bold")
    val_s   = ParagraphStyle("val", parent=styles["Normal"], fontSize=10)

    story = [
        Paragraph(shop_name, h1_s),
        Paragraph("Customer Account Statement", sub_s),
        Paragraph(f"Printed: {datetime.now().strftime('%d-%m-%Y %H:%M')}", sub_s),
        HRFlowable(width="100%", thickness=2,
                   color=colors.HexColor("#2e7d32"), spaceAfter=10),
    ]

    # Customer info
    cust_data = [
        [Paragraph("<b>Name:</b>", lbl_s),    Paragraph(customer.get("name",""), val_s),
         Paragraph("<b>Phone:</b>", lbl_s),   Paragraph(customer.get("phone","") or "—", val_s)],
        [Paragraph("<b>Address:</b>", lbl_s), Paragraph(customer.get("address","") or "—", val_s),
         Paragraph("<b>Notes:</b>", lbl_s),   Paragraph(customer.get("notes","") or "—", val_s)],
    ]
    ct = Table(cust_data, colWidths=[2.5*cm, 7*cm, 2.5*cm, 5*cm])
    ct.setStyle(TableStyle([
        ("TOPPADDING",    (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
    ]))
    story.append(ct)
    story.append(Spacer(1, 0.4*cm))
    story.append(HRFlowable(width="100%", thickness=0.5,
                            color=colors.HexColor("#e2e8f0"), spaceAfter=8))

    # Transactions table
    hdrs = ["Invoice #", "Product", "Qty", "Total", "Paid", "Remaining", "Date"]
    rows = [hdrs]
    for s in sales:
        rem = s.get("remaining_amount", 0) or 0
        rows.append([
            s.get("invoice_number",""),
            s.get("_product",""),
            str(s.get("_qty","")),
            format_currency(s.get("total_amount",0) or 0),
            format_currency(s.get("paid_amount",0) or 0),
            format_currency(rem),
            format_datetime(s.get("sale_date","") or ""),
        ])

    col_w = [3*cm, 5*cm, 1.5*cm, 3*cm, 3*cm, 3*cm, 3.5*cm]
    tbl = Table(rows, colWidths=col_w, repeatRows=1)
    tbl.setStyle(_pdf_table_style(len(hdrs)))
    story.append(tbl)
    story.append(Spacer(1, 0.5*cm))

    # Totals summary
    total_amt  = sum(s.get("total_amount",0) or 0 for s in sales)
    total_paid = sum(s.get("paid_amount",0) or 0 for s in sales)
    total_rem  = sum(s.get("remaining_amount",0) or 0 for s in sales)

    sum_data = [
        ["", "Total Invoiced:", format_currency(total_amt)],
        ["", "Total Paid:",     format_currency(total_paid)],
        ["", "Balance Due:",    format_currency(total_rem)],
    ]
    st = Table(sum_data, colWidths=[9*cm, 4*cm, 4*cm])
    st.setStyle(TableStyle([
        ("ALIGN",     (1,0), (2,-1), "RIGHT"),
        ("FONTNAME",  (1,2), (2,2),  "Helvetica-Bold"),
        ("TEXTCOLOR", (2,2), (2,2),
         colors.HexColor("#dc2626") if total_rem > 0
         else colors.HexColor("#16a34a")),
        ("LINEABOVE", (1,2), (2,2), 1, colors.HexColor("#2e7d32")),
        ("TOPPADDING",    (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
    ]))
    story.append(st)
    doc.build(story)
    return path
